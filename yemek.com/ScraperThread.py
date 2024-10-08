import pandas as pd
import numpy as np
from bs4 import BeautifulSoup as bs
import requests 
import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import re
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import math
import threading
from fractions import Fraction

sections = ["https://yemek.com/tarif/yoresel-tarifler/",
            "https://yemek.com/tarif/kahvaltiliklar/", 
            "https://yemek.com/tarif/zeytinyaglilar/"]


def scroll_down(options):
    # Scroll Down to the end 

    browser = webdriver.Firefox(options=options)
    browser.get("https://yemek.com/tarif/kahvaltiliklar/")

    last_height = browser.execute_script("return document.body.scrollHeight")

    while True:
        browser.execute_script("window.scrollTo(0, document.body.scrollHeight)")

        time.sleep(1)

        new_height = browser.execute_script("return document.body.scrollHeight")

        if new_height == last_height:
            break

        last_height = new_height

    page_source = browser.page_source
    browser.quit()

    return page_source

def get_infinite_content(soup):
    infinite = soup.find("div", attrs={"class":"infinite-scroll-component__outerdiv"})

    infinitecontent = infinite.find_all("div", attrs={"class":"col-md-4 mb-4"})

    return infinitecontent

def set_urls(soup, filename):
    # Open the file in write mode
    infinitecontent = get_infinite_content(soup)
    with open(filename, 'w') as file:
        for content in infinitecontent:
            try:
                # Build the complete URL
                url = "https://yemek.com" + content.a.get("href")
                # Write the URL to the file
                file.write(url + '\n')
            except Exception as e:
                print(f"Error: {e}")
                continue

    print(f"URLs have been written to {filename}.")

def get_urls(filename):
    """
    Reads a text file and returns its contents as a list of lines.

    Parameters:
    - filename: The name of the file to read.

    Returns:
    - A list of lines from the file.
    """
    urls = []
    try:
        with open(filename, 'r') as file:
            urls = [line.strip() for line in file.readlines()]
    except Exception as e:
        print(f"Error reading file {filename}: {e}")
    
    return urls

def find_last_row(excel_file):
    try:
        wb = load_workbook(filename=excel_file)
        ws = wb.active
        # Find the last non-empty row
        last_row = 1
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                last_row += 1
            else:
                break
        last_processed_id = last_row
    except FileNotFoundError:
        last_processed_id = 0

    print(f"Last processed row: {last_processed_id}")
    return last_processed_id

def process_print(processed_count, last_print_time, total_count):
    processed_count += 1
    progress = (processed_count / total_count) * 100

    current_time = time.time()
    if current_time - last_print_time >= 60:  # 60 seconds have passed
        print(f"Progress: {processed_count}/{total_count} ({progress:.2f}%) completed.")
        last_print_time = current_time

    return processed_count

def description_button(driver):
    try:
        # Get button and click it
        descriptionButton = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "ContentRecipe_readMore__5NZAa"))
        )
        driver.execute_script("arguments[0].click();", descriptionButton)
            #python_button.click() #click load more button
    except:
        print("basamadim descriptionButton")
        pass

def description_button_asd(driver):
    # xpath 
    try:
        # Get button and click it
        descriptionButton = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/div/div/div[3]/main/section[2]/div/div[1]/div[3]"))
        )
        descriptionButton.click()
            #python_button.click() #click load more button
    except:
        print("basamadim descriptionButton")
        pass

def nutrition_button(driver):
    try:
        NutritionValuesButton = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "NutritionValuesBox_box__BMJAn"))
        )
        driver.execute_script("arguments[0].click();", NutritionValuesButton)
    except:
        print("basamadim NutritionValuesButton")
        pass

def get_name(soupSelenium):
    # Name 
    name_section = soupSelenium.find("span", attrs={"class": "Breadcrumb_contentTitle__8VL4A"})
    name = name_section.text

    return name

def get_description(soupSelenium):
    try:
        description = soupSelenium.find("article", attrs={"class":"articleContent"})
        description = description.text
    except:
        description=""

    return description

def get_calorie(soupSelenium):
    try:
        # Find the calorie box container
        calorie_section = soupSelenium.find("div", attrs={"class": "ContentRecipe_calorieBox__Y93RU"})

        # Extract each relevant part
        calorie_title = calorie_section.find("span", class_="ContentRecipe_calorieTitle__s_WUP").get_text(strip=True)
        portion_info = calorie_section.find("span", class_="ContentRecipe_portion__PzJSK").get_text(strip=True)
        calorie_amount = calorie_section.find("span", class_="ContentRecipe_calorieAmount__YlH3A").get_text(strip=True)

        # Combine the information into a formatted string
        #calorie_info = f"{calorie_title} {portion_info} - {calorie_amount}"
        calorie_amount = calorie_amount.split('/')[0]
        calorie_info = f"{calorie_amount}"
    except:
        calorie_info = ""

    return calorie_info

def get_nutrition(soupSelenium):
    Protein_value = ""
    Yağ_value = ""
    Karbonhidrat_value = ""

    try: 
        NutritionValues = soupSelenium.find("div", attrs={"class":"NutritionValuesBox_nutritionalValuesBoxWrapper__gNbw9"})
        NutritionValues = NutritionValues.text
        
        # Split the string by lines and iterate over each line
        lines = NutritionValues.splitlines()
        text = '\n'.join(lines)
        return extract_macros(text)

    except Exception as e:
        print(f"Error occurred: {e}")
        NutritionValues = ""
        Protein_value = ""
        Yağ_value = ""
        Karbonhidrat_value = ""

    return Protein_value, Yağ_value, Karbonhidrat_value

def extract_macros(text):
    # Metni küçük harfe çevirip noktalama işaretlerini kaldırıyoruz
    text = text.lower().replace(':', '').replace('gram', '').replace('g', '')
    
    #print(text)
    
    words = text.split()
    
    categories = []
    numbers = []

    if text.find('-')!=-1: # to find out if string has range
        matches = re.findall(r'\b\d+\.?\d*(?=\s*-)', text)
    else:
        matches = re.findall(r'\d+\.?\d*', text)
        if 'için' in text: # to delete first 1
            #print("girdim")
            del matches[0]
    
    numbers = matches
    
    
    for i, word in enumerate(words):
        #print(word)
        if 'protein' in word:
            categories.append("protein")
        elif 'yağ' in word:
            categories.append("yağ")
        elif 'karbonhidrat' in word:
            categories.append("karbonhidrat")

    
    #print("categories: ", categories)
    #print("numbers: ", numbers)
    
    
    return numbers[categories.index('protein')], numbers[categories.index('yağ')], numbers[categories.index('karbonhidrat')]
    

def get_total_time(soupSelenium):
    times = soupSelenium.find_all("div", attrs={"class":"ContentRecipe_recipeDetail__0EBU0"})
    # Initialize a list to accumulate strings (similar to StringBuilder in Java)
    total_time = 0

    # Loop through each recipe detail section
    for each_time in times:
        # Find all <h3> and <span> within this section
        headers = each_time.find_all('h3')
        values = each_time.find_all('span')
        
        # Pair each header with its corresponding value and store them
        for header, value in zip(headers, values):
            header_text = header.get_text()
            value_text = value.get_text()
            
            # Check if it's preparation or cooking time and extract the minutes
            if "HAZIRLAMA SÜRESİ" in header_text:
                # Extract the number of minutes using regex
                preparation_time = int(re.search(r"(\d+)", value_text).group(1))
                total_time += preparation_time
            elif "PİŞİRME SÜRESİ" in header_text:
                # Extract the number of minutes using regex
                cooking_time = int(re.search(r"(\d+)", value_text).group(1))
                total_time += cooking_time

    return total_time

def get_ingredients(soupSelenium):
    ingredient_list = []

    # Find the container that holds the ingredients
    ingredient_section = soupSelenium.find("div", attrs={"class": "Ingredients_ingredients__hk2Pb"})
    # "Ingredients_ingredientList__DhBO1"
    # Extract each ingredient from the list
    ingredient_items = ingredient_section.find_all("li")

    Quantities = []
    Units = []
    Ingredients = []

    # Loop through each ingredient and accumulate the formatted text
    for item in ingredient_items:
        spans = item.find_all("span")  # Find all span elements inside the ingredient
        if len(spans) >= 3:
            quantity = spans[0].get_text().strip()  # 300
            quantity = str(convert_to_float(quantity))
            unit = spans[1].get_text().strip()      # gram
            ingredient = spans[2].get_text().strip() # kıyma
            #print(f"Quantity: {quantity}, Unit: {unit}, Ingredient: {ingredient}")
        #quantity, unit, ingredient = parsingIngredient(ingredient_text)
        Quantities.append(quantity)
        Units.append(unit)
        Ingredients.append(ingredient)
        #ingredient_list.append(ingredient_text)  # Append the ingredient text to the list

    # Join the list into a single string with newline separators (or any delimiter you prefer)
    #ingredients_string = "\n".join(ingredient_list)

    Quantities_str = ', '.join(str(q) for q in Quantities)
    Units_str = ', '.join(Units)
    Ingredients_str = ', '.join(Ingredients)

    return Quantities_str, Units_str, Ingredients_str
    #return Quantities, Units, Ingredients

def convert_to_float(value):
    try:
        # Convert to float if it's a fraction
        if '/' in value:
            return float(Fraction(value))
        else:
            return int(value)
    except ValueError:
        return value

def get_instructions(soupSelenium):
    # Initialize a list to accumulate instructions
    instructions_list = []

    try:
        # Find the section containing the instructions
        instructions_section = soupSelenium.find("div", attrs={"class": "ContentRecipe_instructions__yZRS_"})

        # Find all <li> elements containing individual steps of the instructions
        instruction_items = instructions_section.find_all("li")

        # Loop through each step and accumulate the text
        for index, item in enumerate(instruction_items, start=1):
            step_text = item.get_text(strip=True)  # Get the text inside the <li> and strip whitespace
            instructions_list.append(f"Step {index}: {step_text}")  # Append the formatted step to the list

        # Join the list into a single string with newline separators (or any delimiter you prefer)
        instructions_string = "\n".join(instructions_list)
    except:
        print("error on instructions section")
        instructions_string = ""

    return instructions_string

def write_to_excel(data, excel_file):
    # Create a DataFrame
    df = pd.DataFrame(data)

    # Write the DataFrame to an Excel file
      # Specify your desired file name
    try:
        # Load the existing workbook and select the desired sheet
        wb = load_workbook(filename=excel_file)
        ws = wb['Sheet1']
        
        # Append DataFrame rows to the existing sheet
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)
        
        # Save the updated workbook
        wb.save(excel_file)
        print(f"Data has been appended to {excel_file}")

    # If the file does not exist, create it
    except FileNotFoundError:
        # Write the DataFrame to a new Excel file
        df.to_excel(excel_file, index=False)
        print(f"File not found. A new file has been created: {excel_file}")
    finally:
        # Close the workbook if it's open
        if 'wb' in locals():
            wb.close()

def preparation(options, filename):
    page_source = scroll_down(options)
    # After scrolling, get the fully loaded page's HTML
    soup = bs(page_source, "html.parser")

    set_urls(soup, filename)

def scrape_url(urls, options, start, end, thread_id, last_part):

    driver = None

    for i, url in enumerate(urls[start:end]):
    
        try:

            name = ""
            description = ""
            calorie_info = ""
            NutritionValues = ""
            result_string = ""
            ingredients_string = ""
            instructions_string = ""


            driver = webdriver.Firefox(options=options)
            driver.implicitly_wait(5)
            driver.get(url)

            description_button(driver)
            nutrition_button(driver)

            # Pass to BS4
            soupSelenium = bs(driver.page_source, features="lxml")

            # Values
            name = get_name(soupSelenium)
            name = ' '.join(name.split()[:-1])

            description = get_description(soupSelenium)
            calorie_info = get_calorie(soupSelenium)
            Protein_value, Yağ_value, Karbonhidrat_value = get_nutrition(soupSelenium)
            total_time = get_total_time(soupSelenium)
            Quantities, Units, Ingredients = get_ingredients(soupSelenium)
            instructions_string = get_instructions(soupSelenium)

            # Data to be written to Excel
            data = {
                "Name": [name],
                "Description": [description],
                "Calorie (/kcal)": [calorie_info],
                "Protein":[Protein_value],
                "Fat":[Yağ_value],
                "Carbohydrate":[Karbonhidrat_value],
                "Total_Time": [total_time],
                "Quantities": [Quantities],
                "Units":[Units],
                "Ingredients:":[Ingredients],
                "Instructions": [instructions_string],
            }
            excel_file = f"recipe_data_{last_part}_{thread_id}.xlsx"
            # Write data to Excel
            write_to_excel(data, excel_file)
            
        except Exception as e:
            print(f"Error processing {url}: {e}")
        finally:
            if driver:
                driver.quit()

def prep_for_thread(urls, options, last_part):
    thread_count = 12
    total_url = len(urls)
    #print("total_url: ", total_url)
    thread_list = []
    each_piece = math.floor((total_url-1)/thread_count)
    
    #print("each_piece: ", each_piece)
    for i in range(thread_count):
        start = math.floor(i * each_piece)
        #print("***************************************************start: ", start)
        end = start + each_piece
        th = threading.Thread(target=scrape_url, args=(urls, options, start, end, i, last_part))
        #th.daemon = True # able to kill a thread with Ctrl+C
        thread_list.append(th)

    for thread in thread_list:
        thread.start()

    for thread in thread_list:
        thread.join()

def get_last_part_of_url(url):
    last_part = url.rstrip('/').split('/')[-1]
    return last_part

def main():
    # "2"  "su bardağı" "cart curt"
    options = Options()
    options.add_argument("--headless")
    #options.set_preference('permissions.default.image', 2)
    #options.set_preference('dom.ipc.plugins.enabled.libflashplayer.so', 'false')

    last_part = get_last_part_of_url(sections[2])
    filename = "urls_" + last_part + ".txt"

    # if you need to set urls run following line otherwise dont
    preparation(options, filename)


    urls = get_urls(filename)
    """excel_file = "recipe_data_"+ last_part +".xlsx"
    #excel_file = "thread_recipe_data_"+ last_part +".xlsx"
    last_processed_id = find_last_row(excel_file)"""


    lock = threading.Lock()

    start_time = time.time()
    prep_for_thread(urls, options, last_part)
    end_time = time.time()
    print("Total time: ", str(end_time - start_time), " sec")

if __name__ == "__main__":
    main()